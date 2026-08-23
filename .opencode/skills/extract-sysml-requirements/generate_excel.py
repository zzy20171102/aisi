#!/usr/bin/env python3
"""Generate a styled requirements Excel workbook from a JSON data file.

Usage:
    python generate_excel.py <data.json> [-o output.xlsx]

Input JSON schema:
{
  "title": "项目/文档名称",
  "source": "来源文档文件名",
  "requirements": [
    {
      "id": "REQ-001",
      "name": "需求名称",
      "type": "功能需求",
      "priority": "必须",
      "source": "章节/页码/段落",
      "text": "需求正文",
      "verification": "测试",
      "parent": "REQ-001"   // 可省略
    }
  ]
}
"""

import argparse
import json
import sys
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADERS = ["需求编号", "需求名称", "需求类型", "优先级", "来源", "需求内容", "验证方法", "父需求/关联"]
COL_WIDTHS = [12, 22, 12, 10, 14, 60, 12, 14]

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Microsoft YaHei", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(vertical="top", horizontal="center", wrap_text=True)
BAND_FILL = PatternFill("solid", fgColor="D9E2F3")


def load_data(path: Path) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def build_list_sheet(ws, rows: list[dict]):
    ws.append(HEADERS)
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS[c - 1]
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

    for i, r in enumerate(rows):
        values = [
            r.get("id", ""),
            r.get("name", ""),
            r.get("type", ""),
            r.get("priority", ""),
            r.get("source", ""),
            r.get("text", ""),
            r.get("verification", ""),
            r.get("parent", ""),
        ]
        ws.append(values)
        row = i + 2
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if HEADERS[c - 1] == "需求内容":
                cell.alignment = WRAP_ALIGN
            else:
                cell.alignment = CENTER_ALIGN
            if i % 2 == 1:
                cell.fill = BAND_FILL


def build_stats_sheet(ws, rows: list[dict]):
    title = "需求统计"
    ws.title = title
    ws.append(["统计项", "数量"])
    for c in (1, 2):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12

    ws.append(["需求总数", len(rows)])
    ws.append([])
    ws.append(["按类型"])
    for k, v in Counter(r.get("type", "未分类") for r in rows).most_common():
        ws.append([k, v])
    ws.append([])
    ws.append(["按优先级"])
    for k, v in Counter(r.get("priority", "未指定") for r in rows).most_common():
        ws.append([k, v])
    ws.append([])
    ws.append(["按验证方法"])
    for k, v in Counter(r.get("verification", "未指定") for r in rows).most_common():
        ws.append([k, v])


def main():
    ap = argparse.ArgumentParser(description="Generate requirements xlsx from JSON")
    ap.add_argument("data", type=Path, help="path to JSON data file")
    ap.add_argument("-o", "--output", type=Path, default=None, help="output xlsx path")
    args = ap.parse_args()

    if not args.data.exists():
        sys.exit(f"error: data file not found: {args.data}")

    data = load_data(args.data)
    rows = data.get("requirements", [])
    title = data.get("title", "需求清单")
    source = data.get("source", "")

    out = args.output or (args.data.with_suffix(".xlsx"))
    wb = Workbook()
    ws = wb.active
    ws.title = title if len(title) <= 30 else title[:30]
    build_list_sheet(ws, rows)
    build_stats_sheet(wb.create_sheet(), rows)
    wb.save(out)
    print(f"ok: {out} ({len(rows)} requirements)")


if __name__ == "__main__":
    main()
